from bs4 import BeautifulSoup
import requests
import openpyxl
"""
This will import all the necessary libraries
"""
excel=openpyxl.Workbook()
print(excel.sheetnames)
sheet= excel.active
# Changing the title of the excel
sheet.title='Top Rated Movies'
print(excel.sheetnames)
# creating headings
sheet.append(['Movie Rank','Movie Name','Year Of Release','IMDB Rating'])


try:
    source = requests.get("https://www.imdb.com/chart/top")
    source.raise_for_status()

    soup = BeautifulSoup(source.text,'html.parser')

    movies = soup.find('tbody', class_="lister-list").find_all('tr')

    for movie in movies:
        name= movie.find('td',class_="titleColumn").a.text

        rank= movie.find('td',class_="titleColumn").get_text(strip=True).split(".")[0]

        year= movie.find('td',class_="titleColumn").span.text.strip('()')

        rating= movie.find('td',class_="ratingColumn imdbRating").strong.text

        print(rank,name,year,rating)
        sheet.append([rank,name,year,rating])
        # break


except Exception as e:
    print(e)
# Saving the excel file and giving its name
excel.save('IMDB Movie Rating.xlsx')